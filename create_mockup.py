import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def create_excel_mockup():
    output_dir = r"d:\Joshi\task2_eda"
    output_path = os.path.join(output_dir, "dashboard_mockup.xlsx")
    
    os.makedirs(output_dir, exist_ok=True)
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: Executive Mockup
    ws1 = wb.active
    ws1.title = "Dashboard Mockup"
    ws1.views.sheetView[0].showGridLines = True
    
    # Custom colors
    teal_fill = PatternFill(start_color="0F4C5C", end_color="0F4C5C", fill_type="solid")
    accent_fill = PatternFill(start_color="E36414", end_color="E36414", fill_type="solid")
    gray_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Fonts
    title_font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    section_font = Font(name="Arial", size=12, bold=True, color="1D3557")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    label_font = Font(name="Arial", size=9, bold=True, color="64748B")
    value_font = Font(name="Arial", size=18, bold=True, color="0F4C5C")
    sub_font = Font(name="Arial", size=8, italic=True, color="E36414")
    bold_cell_font = Font(name="Arial", size=10, bold=True)
    
    # Borders
    thin_border_side = Side(style='thin', color='DDDDDD')
    thick_border_side = Side(style='medium', color='0F4C5C')
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    top_thick_border = Border(top=thick_border_side, left=thin_border_side, right=thin_border_side, bottom=thin_border_side)
    
    # Header Banner
    ws1.merge_cells("B2:J3")
    for r in range(2, 4):
        for c in range(2, 11):
            cell = ws1.cell(row=r, column=c)
            cell.fill = teal_fill
            
    header_cell = ws1["B2"]
    header_cell.value = "ONLINE RETAIL EXECUTIVE KPI DASHBOARD - WIREFRAME MOCKUP"
    header_cell.font = title_font
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # KPI 1: Net Revenue
    ws1.merge_cells("B5:C6")
    for r in range(5, 7):
        for c in range(2, 4):
            ws1.cell(row=r, column=c).border = cell_border
    ws1["B5"].value = "TOTAL NET REVENUE\n\n£9,742,030.56\n(Net Sales value)"
    ws1["B5"].font = bold_cell_font
    ws1["B5"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws1["B5"].border = top_thick_border
    
    # KPI 2: Total Orders
    ws1.merge_cells("D5:E6")
    for r in range(5, 7):
        for c in range(4, 6):
            ws1.cell(row=r, column=c).border = cell_border
    ws1["D5"].value = "COMPLETED ORDERS\n\n19,960\n(Shipped volume)"
    ws1["D5"].font = bold_cell_font
    ws1["D5"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws1["D5"].border = top_thick_border
    
    # KPI 3: AOV
    ws1.merge_cells("F5:G6")
    for r in range(5, 7):
        for c in range(6, 8):
            ws1.cell(row=r, column=c).border = cell_border
    ws1["F5"].value = "AVERAGE ORDER VALUE\n\n£488.08\n(Avg basket spend)"
    ws1["F5"].font = bold_cell_font
    ws1["F5"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws1["F5"].border = top_thick_border
    
    # KPI 4: Active Customers
    ws1.merge_cells("H5:I6")
    for r in range(5, 7):
        for c in range(8, 10):
            ws1.cell(row=r, column=c).border = cell_border
    ws1["H5"].value = "ACTIVE BUYERS\n\n4,321\n(Registered & guests)"
    ws1["H5"].font = bold_cell_font
    ws1["H5"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws1["H5"].border = top_thick_border
    
    # Section: Monthly Trends Data Table Mockup
    ws1["B8"] = "1. Monthly Sales Trajectory Data Table"
    ws1["B8"].font = section_font
    
    headers = ["Month", "Revenue (£)", "Active Customers", "Total Orders"]
    for idx, h in enumerate(headers):
        cell = ws1.cell(row=9, column=2+idx)
        cell.value = h
        cell.fill = teal_fill
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        
    mockdata = [
        ("2011-09", 1017596.68, 1395, 2170),
        ("2011-10", 1069368.23, 1564, 2402),
        ("2011-11", 1456145.80, 1835, 3210),
    ]
    
    for row_idx, data in enumerate(mockdata):
        for col_idx, val in enumerate(data):
            cell = ws1.cell(row=10+row_idx, column=2+col_idx)
            cell.value = val
            cell.border = cell_border
            if col_idx == 1:
                cell.number_format = '"£"#,##0.00'
            elif col_idx in [2, 3]:
                cell.number_format = '#,##0'
                
    # Summary info box mockup
    ws1.merge_cells("G9:I12")
    for r in range(9, 13):
        for c in range(7, 10):
            ws1.cell(row=r, column=c).border = cell_border
    ws1["G9"].value = "DASHBOARD TAKEAWAYS:\n\n• Sales spike 2.9x in November vs. Spring base.\n• Midday hours (10 AM - 3 PM) account for 70% of transactions.\n• International AOV is 1.7x higher than domestic UK AOV."
    ws1["G9"].font = Font(name="Arial", size=9, italic=True)
    ws1["G9"].alignment = Alignment(wrap_text=True, vertical="top")
    ws1["G9"].fill = gray_fill
    
    # Format column widths
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Save workbook
    wb.save(output_path)
    print(f"Excel mockup successfully saved to {output_path}")

if __name__ == "__main__":
    create_excel_mockup()